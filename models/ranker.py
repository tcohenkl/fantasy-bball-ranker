# Blends the stat model baseline (65%) with the intuition model (35%).
# Top 60 players by fantasy PPG enter a round-robin tournament; each pair
# is scored by the trained logistic classifier, then merged with the stat rank.

import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import numpy as np

sys.path.insert(0, str(Path(__file__).parent.parent))
from db import CURRENT_SEASON, OUTPUT_DIR, get_conn
from models.intuition_model import (
    _delta,
    _load_artifact,
    _player_stats_map,
    _usg_pg,
    get_model_confidence,
)

_STAT_TOP_N  = 60
_FINAL_TOP_N = 40
_BLEND_ALPHA = 0.35  # weight given to intuition score vs stat rank


def _tournament_scores(players: list[dict], artifact: dict) -> list[float]:
    """Run every pair through the classifier; accumulate win probabilities."""
    clf, scaler = artifact["clf"], artifact["scaler"]
    n      = len(players)
    scores = [0.0] * n
    for i in range(n):
        for j in range(i + 1, n):
            X      = scaler.transform([_delta(players[i], players[j])])
            prob_i = float(clf.predict_proba(X)[0][1])
            scores[i] += prob_i
            scores[j] += (1.0 - prob_i)
    return scores


def _build_ranking(season: str, top_n: int, player_filter: list[str] | None = None) -> tuple[list[dict], str]:
    """
    Returns (ranked_rows, confidence).
    Each row is the player dict with final_rank, stat_rank, and movement keys added.
    """
    stats_map = _player_stats_map(season)
    if not stats_map:
        return [], "NO DATA"

    if player_filter:
        name_lower = [n.strip().lower() for n in player_filter]
        filtered: dict[int, dict] = {}
        for pid, p in stats_map.items():
            full     = (p.get("name") or "").lower()
            parts    = full.split()
            initials = "".join(w[0] for w in full.replace("-", " ").split())
            first    = parts[0] if parts else full
            last     = parts[-1] if parts else full
            for q in name_lower:
                if (q in full or full in q or q == initials
                        or q == first or q == last
                        or (len(q) >= 4 and q in full.replace("-", ""))):
                    filtered[pid] = p
                    break
        stats_map = filtered

    min_gp = 10 if not player_filter else 0
    by_fpg = sorted(
        (p for p in stats_map.values() if (p.get("gp") or 0) >= min_gp),
        key=lambda p: float(p.get("fantasy_ppg") or 0),
        reverse=True,
    )[:_STAT_TOP_N]

    if not by_fpg:
        return [], "NO DATA"

    stat_ranks = {p["player_id"]: i + 1 for i, p in enumerate(by_fpg)}
    artifact   = _load_artifact()
    confidence = get_model_confidence()

    if artifact:
        t_scores = _tournament_scores(by_fpg, artifact)
        max_t    = max(t_scores) or 1.0
        t_norm   = [s / max_t for s in t_scores]
        alpha    = _BLEND_ALPHA
        n        = len(by_fpg)
        blended  = []
        for i, player in enumerate(by_fpg):
            blended.append(((1 - alpha) * (1.0 - i / n) + alpha * t_norm[i], player))
        final = [p for _, p in sorted(blended, key=lambda t: t[0], reverse=True)]
    else:
        final      = by_fpg
        confidence = "NO MODEL"

    rows = []
    for final_rank, player in enumerate(final[:top_n], 1):
        pid       = player["player_id"]
        stat_rank = stat_ranks.get(pid, 0)
        diff      = (stat_rank - final_rank) if stat_rank else 0
        movement  = (f"+{diff}" if diff > 0 else str(diff)) if diff != 0 else ""
        row       = dict(player)
        row["final_rank"] = final_rank
        row["stat_rank"]  = stat_rank
        row["movement"]   = movement
        rows.append(row)

    return rows, confidence


def _export_rankings_excel(rows: list[dict], season: str, confidence: str, label: str = "") -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Run: pip install openpyxl")

    HEADER_FILL = PatternFill("solid", fgColor="0D1B2A")
    HEADER_FONT = Font(bold=True, color="F97316")
    ALT_FILL    = PatternFill("solid", fgColor="1A1A2E")
    OUT_FILL    = PatternFill("solid", fgColor="7F1D1D")
    DTD_FILL    = PatternFill("solid", fgColor="78350F")
    WHITE_FONT  = Font(color="FFFFFF")
    DIM_FONT    = Font(color="AAAAAA")
    CENTER      = Alignment(horizontal="center")

    COLS = [
        ("Rank",    5),  ("Player",  22),  ("Pos",   5),  ("Team",   7),
        ("W-L",     7),  ("Seed",    6),   ("Inj",   6),
        ("Fan PPG", 9),  ("PTS",     7),   ("REB",   7),  ("AST",    7),
        ("STL",     7),  ("BLK",     7),   ("TOV",   7),
        ("3PM",     7),  ("3PA",     7),   ("3P%",   7),
        ("FGM",     7),  ("FGA",     7),   ("FG%",   7),
        ("FTM",     7),  ("FTA",     7),   ("FT%",   7),
        ("USG%",    7),  ("MIN",     7),   ("GP",    5),  ("Prev GP", 8),
        ("Stat Rk", 8),  ("Move",    6),
    ]

    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = season
    ws.sheet_properties.tabColor = "F97316"

    title = f"Fantasy Rankings -- {season}  |  Confidence: {confidence}  |  {date.today()}"
    if label:
        title += f"  |  {label}"
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    tc = ws["A1"]
    tc.value     = title
    tc.font      = Font(bold=True, color="F97316", size=12)
    tc.fill      = PatternFill("solid", fgColor="0D1B2A")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, (col_name, col_w) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=col_name)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = col_w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}2"

    for row in rows:
        inj     = str(row.get("injury_status") or "ACTIVE").upper()
        fgm     = float(row.get("fgm")  or 0)
        fga     = float(row.get("fga")  or 0)
        ftm     = float(row.get("ftm")  or 0)
        fta     = float(row.get("fta")  or 0)
        fg3m    = float(row.get("fg3m") or 0)
        fg3a    = float(row.get("fg3a") or 0)
        mpg     = float(row.get("min")  or 0)
        usg     = round((_usg_pg(row) / mpg * 36) if mpg else 0, 1)
        fg_pct  = round(fgm / fga, 3) if fga else 0
        ft_pct  = round(ftm / fta, 3) if fta else 0
        fg3_pct = round(fg3m / fg3a, 3) if fg3a else 0
        abbr    = str(row.get("team_abbr") or "")
        wins    = int(row.get("team_wins")   or 0)
        losses  = int(row.get("team_losses") or 0)
        seed    = int(row.get("team_seed")   or 0)
        conf    = str(row.get("team_conf")   or "")
        seed_str = f"#{seed} {conf[:1]}" if seed else ""
        prev_gp = int(row.get("gp_prev_season") or 0)

        values = [
            row["final_rank"],
            row.get("name", ""),
            row.get("position") or "?",
            abbr,
            f"{wins}-{losses}" if abbr else "",
            seed_str,
            inj,
            round(float(row.get("fantasy_ppg") or 0), 1),
            round(float(row.get("pts") or 0), 1),
            round(float(row.get("reb") or 0), 1),
            round(float(row.get("ast") or 0), 1),
            round(float(row.get("stl") or 0), 1),
            round(float(row.get("blk") or 0), 1),
            round(float(row.get("tov") or 0), 1),
            round(fg3m, 1),
            round(fg3a, 1),
            fg3_pct,
            round(fgm, 1),
            round(fga, 1),
            fg_pct,
            round(ftm, 1),
            round(fta, 1),
            ft_pct,
            usg,
            round(mpg, 1),
            int(row.get("gp") or 0),
            prev_gp or "",
            f"#{row['stat_rank']}" if row["stat_rank"] else "",
            row.get("movement", ""),
        ]

        rn   = row["final_rank"] + 2
        fill = ALT_FILL if row["final_rank"] % 2 == 0 else PatternFill("solid", fgColor="0F0F1A")
        if inj in ("OUT", "INJURED_RESERVE"):
            fill = OUT_FILL
        elif inj in ("QUESTIONABLE", "DAY_TO_DAY", "DOUBTFUL", "PROBABLE"):
            fill = DTD_FILL

        for ci, val in enumerate(values, 1):
            c            = ws.cell(row=rn, column=ci, value=val)
            c.fill       = fill
            c.font       = WHITE_FONT if ci == 2 else DIM_FONT
            c.alignment  = CENTER if ci != 2 else Alignment(horizontal="left")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUTPUT_DIR / "rankings.xlsx"
    wb.save(out)
    return out


def generate_rankings(season: str = CURRENT_SEASON, top_n: int = _FINAL_TOP_N) -> list[dict]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows, confidence = _build_ranking(season, top_n)
    if not rows:
        print("No player data. Run `python main.py ingest` first.")
        return []

    _print_table(rows, season, confidence)
    path = _export_rankings_excel(rows, season, confidence)
    print(f"\n[Saved -> {path.name}]")
    return rows


def rank_players(season: str, player_names: list[str] | None = None) -> list[dict]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows, confidence = _build_ranking(season, _STAT_TOP_N, player_names)
    if not rows:
        print(f"No data for {season}. Run `python main.py ingest` first.")
        return []

    label = f"Filter: {', '.join(player_names)}" if player_names else ""
    _print_table(rows, season, confidence, label=label)
    path = _export_rankings_excel(rows, season, confidence, label=label)
    print(f"\n[Saved -> {path.name}]")
    return rows


def _print_table(rows: list[dict], season: str, confidence: str, label: str = "") -> None:
    today = date.today().isoformat()
    hdr   = f"Fantasy Rankings -- {season}  [{confidence}]  {today}"
    if label:
        hdr += f"  |  {label}"
    print(f"\n{hdr}")
    print("=" * len(hdr))
    print()

    fmt = (f"{'#':<4} {'Player':<24} {'Pos':<4} {'Team':<13} {'Inj':<5} "
           f"{'FPPG':<7} {'PTS':<6} {'REB':<5} {'AST':<5} "
           f"{'STL':<5} {'BLK':<5} {'TOV':<5} {'3PM':<5} "
           f"{'FG%':<6} {'FT%':<6} {'USG%':<6} {'GP':<4} {'Prev':<5} {'Mv'}")
    print(fmt)
    print("-" * len(fmt))

    for r in rows:
        fgm     = float(r.get("fgm") or 0)
        fga     = float(r.get("fga") or 0)
        ftm     = float(r.get("ftm") or 0)
        fta     = float(r.get("fta") or 0)
        mpg     = float(r.get("min") or 0)
        fg_pct  = f"{fgm/fga*100:.0f}%" if fga else "--"
        ft_pct  = f"{ftm/fta*100:.0f}%" if fta else "--"
        usg     = f"{_usg_pg(r)/mpg*36:.0f}%" if mpg else "--"
        abbr    = str(r.get("team_abbr") or "")
        wins    = int(r.get("team_wins") or 0)
        losses  = int(r.get("team_losses") or 0)
        team    = f"{abbr} {wins}-{losses}" if abbr else "--"
        inj     = str(r.get("injury_status") or "")
        inj_s   = "OUT" if "OUT" in inj else ("Q" if any(x in inj for x in ("QUEST","DAY","DOUBT")) else "")
        prev    = int(r.get("gp_prev_season") or 0)

        print(
            f"{r['final_rank']:<4} {str(r.get('name',''))[:23]:<24} "
            f"{str(r.get('position') or '?'):<4} {team:<13} {inj_s:<5} "
            f"{float(r.get('fantasy_ppg') or 0):<7.1f} "
            f"{float(r.get('pts') or 0):<6.1f} {float(r.get('reb') or 0):<5.1f} "
            f"{float(r.get('ast') or 0):<5.1f} {float(r.get('stl') or 0):<5.1f} "
            f"{float(r.get('blk') or 0):<5.1f} {float(r.get('tov') or 0):<5.1f} "
            f"{float(r.get('fg3m') or 0):<5.1f} {fg_pct:<6} {ft_pct:<6} {usg:<6} "
            f"{int(r.get('gp') or 0):<4} {prev or '--':<5} {r.get('movement','')}"
        )


if __name__ == "__main__":
    generate_rankings()
