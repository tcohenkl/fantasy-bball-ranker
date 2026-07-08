"""
Export player stats to Excel.  One sheet per season, top 100 by fantasy PPG.
Columns: Rank, Player, Pos, Injury, GP, Fantasy PPG, PTS, REB, AST, STL, BLK,
         TOV, FGM, FGA, FG%, FTM, FTA, FT%, 3PM, Rookie, Prev GP
"""

import sqlite3
from pathlib import Path

OUTPUT_DIR = Path(__file__).parent.parent / "outputs"


def export_to_excel(db_path: Path, seasons: list[str] | None = None) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Run: pip install openpyxl")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    if seasons is None:
        seasons = [
            r[0] for r in conn.execute(
                "SELECT DISTINCT season FROM player_stats ORDER BY season DESC"
            ).fetchall()
        ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT   = Font(bold=True, color="FFFFFF")
    ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")
    INJURY_FILL   = PatternFill("solid", fgColor="FFD700")
    OUT_FILL      = PatternFill("solid", fgColor="FF6B6B")
    CENTER        = Alignment(horizontal="center")

    COLS = [
        ("Rank",       6),   ("Player",     24),  ("Pos",        5),
        ("Injury",     10),  ("GP",          5),  ("Fan PPG",    9),
        ("PTS",         7),  ("REB",         7),  ("AST",        7),
        ("STL",         7),  ("BLK",         7),  ("TOV",        7),
        ("FGM",         7),  ("FGA",         7),  ("FG%",        7),
        ("FTM",         7),  ("FTA",         7),  ("FT%",        7),
        ("3PM",         7),  ("Prev GP",     8),
    ]

    for season in seasons:
        ws = wb.create_sheet(title=season)

        # Header row
        for col_idx, (col_name, col_width) in enumerate(COLS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        ws.freeze_panes = "A2"

        rows = conn.execute(
            """
            SELECT
                p.name, p.position, p.injury_status,
                ps.gp, ps.pts, ps.reb, ps.ast, ps.stl, ps.blk, ps.tov,
                ps.fgm, ps.fga, ps.ftm, ps.fta, ps.fg3m,
                ps.is_rookie, ps.gp_prev_season,
                fs.fantasy_ppg
            FROM player_stats ps
            JOIN players p         ON p.player_id = ps.player_id
            JOIN fantasy_scores fs ON fs.player_id = ps.player_id AND fs.season = ps.season
            WHERE ps.season = ?
            ORDER BY fs.fantasy_ppg DESC
            LIMIT 100
            """,
            (season,),
        ).fetchall()

        for rank, r in enumerate(rows, 1):
            fgm  = r["fgm"] or 0
            fga  = r["fga"] or 0
            ftm  = r["ftm"] or 0
            fta  = r["fta"] or 0
            fg_pct = round(fgm / fga, 3) if fga else 0
            ft_pct = round(ftm / fta, 3) if fta else 0
            inj    = (r["injury_status"] or "ACTIVE").upper()

            values = [
                rank,
                r["name"],
                r["position"] or "?",
                inj,
                r["gp"],
                round(r["fantasy_ppg"], 1),
                round(r["pts"],  1),
                round(r["reb"],  1),
                round(r["ast"],  1),
                round(r["stl"],  1),
                round(r["blk"],  1),
                round(r["tov"],  1),
                round(fgm,  1),
                round(fga,  1),
                fg_pct,
                round(ftm,  1),
                round(fta,  1),
                ft_pct,
                round(r["fg3m"] or 0, 1),
                r["gp_prev_season"] or "",
            ]

            row_num = rank + 1
            fill    = ALT_FILL if rank % 2 == 0 else None
            if inj in ("OUT", "INJURED_RESERVE"):
                fill = OUT_FILL
            elif inj in ("QUESTIONABLE", "DAY_TO_DAY", "DOUBTFUL", "PROBABLE"):
                fill = INJURY_FILL

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                if fill:
                    cell.fill = fill
                if col_idx == 1 or col_idx >= 3:  # Rank + everything after Player
                    cell.alignment = CENTER

        # Freeze header and auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    conn.close()

    OUTPUT_DIR.mkdir(exist_ok=True)
    out_path = OUTPUT_DIR / "player_rankings.xlsx"
    wb.save(out_path)
    return out_path
